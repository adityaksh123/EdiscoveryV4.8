from selenium import webdriver
import time
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains, Keys
import random
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
# serv_object = Service(executable_path="/Users/adityakshirsagar/chromedriver-mac-x64/chromedriver")
# driver = webdriver.Chrome(service=serv_object)
# act = ActionChains(driver)



def initialise_driver():
    serv_object = Service(executable_path="/Users/adityakshirsagar/chromedriver-mac-x64/chromedriver")
    driver = webdriver.Chrome(service=serv_object)
    driver.get("https://qa.servient.net/ng/#/login")
    driver.maximize_window()
    return driver

def login(driver,username, password):
    driver.get("https://qa.servient.net/ng/#/login")
    driver.maximize_window()
    driver.find_element(By.ID,'mat-input-0').send_keys(username)   #Login
    driver.find_element(By.ID,'mat-input-1').send_keys(password)
    driver.find_element(By.XPATH,'//*[@id="cardAction"]/button').click()


def subscriber(driver,subscriber_name):
    noRows = len(driver.find_elements(By.XPATH, '/html/body/app-root/ser-select-subscriber/div/div[2]/ser-subscriber-grid/div/div/kendo-grid/div/kendo-grid-list/div/div[1]/table/tbody/tr'))
    for r in range(1, noRows + 1):
        sub_list = driver.find_elements(By.XPATH, f'/html/body/app-root/ser-select-subscriber/div/div[2]/ser-subscriber-grid/div/div/kendo-grid/div/kendo-grid-list/div/div[1]/table/tbody/tr[{r}]/td[1]')
        for subscriber in sub_list:
            if subscriber.text == subscriber_name:
                subscriber.click()
                break
    time.sleep(5)


def case (driver,case_name):
    driver.find_element(By.XPATH, '/html/body/app-root/ser-app-list/div/div[2]/div[1]/div[5]/div/div[1]').click()
    time.sleep(5)
    noCase = len(driver.find_elements(By.XPATH, '/html/body/app-root/ser-switchboard/ser-main-app/mat-sidenav-container/mat-sidenav-content/div/main/ser-case-list/div/div/kendo-grid/div/kendo-grid-list/div/div[1]/table/tbody/tr'))
    for c in range(1, noCase + 1):
        case_list = driver.find_elements(By.XPATH, f'/html/body/app-root/ser-switchboard/ser-main-app/mat-sidenav-container/mat-sidenav-content/div/main/ser-case-list/div/div/kendo-grid/div/kendo-grid-list/div/div[1]/table/tbody/tr[{c}]/td[1]')
        driver.implicitly_wait(5)
        for case in case_list:
            if case.text == case_name:
                case.click()
                break
    time.sleep(5)

# search_bar = driver.find_element(By.CLASS_NAME,'cdk-textarea-autosize search-bar ser-body-1 default ng-tns-c1052-23 ng-pristine ng-valid ng-touched').send_keys(data("Sheet 1"))
# time.sleep(10)


# def data (driver,sheet_name):


def main():
    driver = initialise_driver()
    login(driver, "automation2@servient.com", "Servient@123")
    subscriber(driver, "Tomcat 9")
    case(driver, "15OctSmoke")

import openpyxl
file = '/Users/adityakshirsagar/Downloads/search.xlsx'
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet 1"]
rows = sheet.max_row
columns = sheet.max_column
for r in range(4, rows + 1):
    text = sheet.cell(r, 2).value
    print(text, end='')
    print()
